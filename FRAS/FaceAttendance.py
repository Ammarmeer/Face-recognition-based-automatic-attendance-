from sys import path
import csv
from tkinter import*
from tkinter import ttk
from PIL import Image,ImageTk
import os
import mysql.connector
import cv2
from tkinter import filedialog as fd
import numpy as np
import openpyxl
from openpyxl import workbook,load_workbook
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from datetime import date
import datetime
from datetime import datetime
from tkinter import messagebox
import time
import pkg_resources
import face_recognition
#haar=pkg_resources.resource_filename('cv2','data/haarcascade_frontalface_default.xml')
directory="C:\\Users\\AMMAR MEER\\Desktop\\FRAS\\"
# from PIL import ImageGrab

class facerecognitionsystem:
    def __init__(self, FR_wind):
        #Setting Up window form
        self.FR_wind=FR_wind
        self.FR_wind.geometry("1368x768+0+0")
        self.FR_wind.title("Mark Attendance")

        img=Image.open(r"Images_GUI\bg2.jpg")
        img=img.resize((1368,768),Image.ANTIALIAS)
        self.photoimg=ImageTk.PhotoImage(img)
        bck_lbl=Label(self.FR_wind,image=self.photoimg)
        bck_lbl.place(x=0,y=0,width=1368, height=768)

        self.var_subject=StringVar()

        # bck_lbl=Label(self.FR_wind,text="Hellp")
        # bck_lbl.place(x=0,y=0,width=1300, height=700)
        #startMarking
        btnImg2=Image.open(r"Images_GUI\f_det.jpg")
        btnImg2=btnImg2.resize((180,180),Image.ANTIALIAS)
        self.photoimg2=ImageTk.PhotoImage(btnImg2)
        b1=Button(bck_lbl,image=self.photoimg2,cursor="hand2",command=self.facerecognize)
        b1.place(x=485,y=260,width=180,height=180)

        std_b1_1 = Button(bck_lbl,text="Start Attendance",cursor="hand2",font=("tahoma",15,"bold"),bg="white",fg="navyblue")
        std_b1_1.place(x=485,y=430,width=180,height=45)
        
        
        std_img_btn1=Image.open(r"Images_GUI\save-icon.png")
        std_img_btn1=std_img_btn1.resize((180,180),Image.ANTIALIAS)
        self.std_img2=ImageTk.PhotoImage(std_img_btn1)

        std_b2 = Button(bck_lbl,image=self.std_img2,cursor="hand2",command=self.txttolist)
        std_b2.place(x=685,y=260,width=180,height=180)


        std_b1_2 = Button(bck_lbl,text="Save Attendance",cursor="hand2",font=("tahoma",15,"bold"),bg="white",fg="navyblue")
        std_b1_2.place(x=685,y=430,width=180,height=45)

        left_frame = LabelFrame(bck_lbl,bd=2,bg="white",relief=RIDGE,text="Attendance Sheets",font=("verdana",12,"bold"),fg="navyblue")
        left_frame.place(x=580,y=10,width=200,height=80)

        sub_combo=ttk.Combobox(left_frame,textvariable=self.var_subject,width=15,font=("verdana",12,"bold"),state="readonly")
        sub_combo["values"]=("Select Subject","Dummy Attendance","SE Economics","Arabic","NLP","Topics in SE")
        sub_combo.current(0)
        sub_combo.grid(row=0,column=1,padx=5,pady=15,sticky=W)

    def browse(self):
        filename = fd.askopenfilename()
        #messagebox.showinfo("path",filename)
        return filename


    def save_raw_Attendance(self,n):
        today = str(date.today())
        txt="raw_"+today+".txt"
        filepath = directory
        file=filepath+txt
        with open(filepath+"raw_"+today+".txt", "a+",newline="\n") as fp:
                myDatalist=fp.readlines()
                name_list=[]
                for line in myDatalist:
                    entry=line.split((","))
                    name_list.append(entry[0])
                    ##print(name_list)
                if((n not in name_list)):
                    now=datetime.now()
                    d1=now.strftime("%d/%m/%Y")
                    dtString=now.strftime("%H:%M:%S")
                    fp.writelines(n+ '\n')

        def remove_repeated(file):
            openFile = open(file, "r") 
            writeFile = open("text.txt", "w") #bug
            #Store traversed lines
            tmp = set() 
            for txtLine in openFile: 
            #Check new line
                if txtLine not in tmp: 
                    writeFile.write(txtLine) 
            #Add new traversed line to tmp 
                    tmp.add(txtLine)         
            openFile.close() 
            writeFile.close()
        remove_repeated(file)
    

    def markA(self,namelist):
        sub=self.var_subject.get()
        if sub!="Select Subject":

            filename = fd.askopenfilename()
            #messagebox.showinfo("path",filename)
            column=0
            today = date.today()
            td=str(today)
            datedata=""
            datad=0
            column=0
            td=str(today)
            wb=load_workbook(filename)
            ws=wb.active
            sheets = wb.sheetnames
            mont=str(today.month)
            if mont=="9":
                ws = wb[sheets[0]]
            elif mont=="10":
                ws = wb[sheets[1]]
                print(ws)
            elif mont=="11":
                ws = wb[sheets[2]]
            elif mont=="12":
                ws = wb[sheets[3]]

            dateindex=0
            for dtrows in range(8,9):
                for dtcol in range(3,12):
                    char=get_column_letter(dtcol)
                    val=ws[char+str(dtrows)].value
                    val=str(val)
                    datedata=val[0:10]
                    if td==datedata:
                        dateindex=dtcol
                        break
            if td==datedata:
                column=dateindex
            else:
                messagebox.showerror("Hold on","No class today!!")
            rowdata=0
            for rows in range(10,41):
                for col in range(2,3):
                    char=get_column_letter(col)
                    val=ws[char+str(rows)].value
                    #print(val)
                    for i in namelist:
                        if i==val:
                            row=rowdata+rows
                            print("line 161 working")
                            #print(row)
                            c1 = ws.cell(row, column)
                            c1.value = "P" 
            
            wb.save("C:/Users/AMMAR MEER/Desktop/FRAS/Attendance/sheet_"+sub+".xlsx")
            messagebox.showinfo("Saved","Attendance Successfully saved!")
        else:
            messagebox.showerror("Error","Please Select the subject first!")

    def txttolist(self):
        my_file = open("text.txt", "r")
        # reading the file
        data = my_file.read()
        # replacing end splitting the text 
        # when newline ('\n') is seen.
        data_into_list = data.split("\n")
        correctlist=data_into_list.pop()
        my_file.close()
        self.markA(data_into_list)

    
            # with open('Attendance.csv','r+') as f:
            #     myDataList = f.readlines()
            #     nameList = []
            #     for line in myDataList:
            #         entry = line.split(',')
            #         nameList.append(entry[0])
            #     if name not in nameList:
            #         now = datetime.now()
            #         dtString = now.strftime('%H:%M:%S')
            #         f.writelines(f'\n{name},{dtString},present')
    
    def facerecognize(self):
            path = 'ImagesOfFaces'
            images = []
            classNames = []
            myList = os.listdir(path)
            print(myList)
            for cl in myList:
                curImg = cv2.imread(f'{path}/{cl}')
                images.append(curImg)
                classNames.append(os.path.splitext(cl)[0])
            print(classNames)

            def findEncodings(images):
                encodeList = []
                for img in images:
                    img = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)
                    encode = face_recognition.face_encodings(img)[0]
                    encodeList.append(encode)
                return encodeList
            encodeListKnown = findEncodings(images)
            print('Encoding Complete')

            cap = cv2.VideoCapture(0)
                
            while True:
                success, img = cap.read()
                #img = captureScreen()
                imgS = cv2.resize(img,(0,0),None,0.25,0.25)
                imgS = cv2.cvtColor(imgS, cv2.COLOR_BGR2RGB)
                    
                facesCurFrame = face_recognition.face_locations(imgS)
                encodesCurFrame = face_recognition.face_encodings(imgS,facesCurFrame)
                    
                for encodeFace,faceLoc in zip(encodesCurFrame,facesCurFrame):
                    matches = face_recognition.compare_faces(encodeListKnown,encodeFace)
                    faceDis = face_recognition.face_distance(encodeListKnown,encodeFace)
                    #print(faceDis)
                    matchIndex = np.argmin(faceDis)
                    if faceDis[matchIndex]< 0.50:
                        name = classNames[matchIndex]
                        self.save_raw_Attendance(name)

                    else: name = 'Unknown'
                    #print(name)
                    y1,x2,y2,x1 = faceLoc
                    y1, x2, y2, x1 = y1*4,x2*4,y2*4,x1*4
                    cv2.rectangle(img,(x1,y1),(x2,y2),(0,0,255),2)
                    cv2.rectangle(img,(x1,y2-35),(x2,y2),(0,0,255),cv2.FILLED)
                    cv2.putText(img,name,(x1+6,y2-6),cv2.FONT_HERSHEY_COMPLEX,0.5,(255,255,255),1)

                cv2.imshow('Webcam',img)
                if cv2.waitKey(10) == ord('q'):           
                    cap.release()
                    cv2.destroyAllWindows()


if __name__ =="__main__":
    FR_wind=Tk()
    obj=facerecognitionsystem(FR_wind)
    FR_wind.mainloop()  