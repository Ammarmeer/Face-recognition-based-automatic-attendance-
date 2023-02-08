# import re
from sys import path
import csv
from tkinter import*
from tkinter import ttk
from PIL import Image,ImageTk
import os
import mysql.connector
import cv2
import numpy as np
import openpyxl
from openpyxl import workbook,load_workbook
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from datetime import date
import datetime
from datetime import datetime
from tkinter import messagebox
import time
import pkg_resources
import face_recognition
haar=pkg_resources.resource_filename('cv2','data/haarcascade_frontalface_default.xml')
directory="C:\\Users\\AMMAR MEER\\Desktop\\Python-FYP-Face-Recognition-Attendence-System-master\\"

class FaceDetect:
    def __init__(self,root):
        self.root=root
        self.root.geometry("1366x768+0+0")
        self.root.title("Face Recognition Pannel")

        # This part is image labels setting start 
        # first header image  
        img=Image.open(r"Images_GUI\banner.jpg")
        img=img.resize((1366,130),Image.ANTIALIAS)
        self.photoimg=ImageTk.PhotoImage(img)

        # set image as lable
        f_lb1 = Label(self.root,image=self.photoimg)
        f_lb1.place(x=0,y=0,width=1366,height=130)

        # backgorund image 
        bg1=Image.open(r"Images_GUI\bg2.jpg")
        bg1=bg1.resize((1366,768),Image.ANTIALIAS)
        self.photobg1=ImageTk.PhotoImage(bg1)

        # set image as lable
        bg_img = Label(self.root,image=self.photobg1)
        bg_img.place(x=0,y=130,width=1366,height=768)


        #title section
        title_lb1 = Label(bg_img,text="Welcome to Face Recognition Panel",font=("verdana",30,"bold"),bg="white",fg="navyblue")
        title_lb1.place(x=0,y=0,width=1366,height=45)

        # Create buttons below the section 
        # ------------------------------------------------------------------------------------------------------------------- 
        # Training button 1
        std_img_btn=Image.open(r"Images_GUI\f_det.jpg")
        std_img_btn=std_img_btn.resize((180,180),Image.ANTIALIAS)
        self.std_img1=ImageTk.PhotoImage(std_img_btn)

        std_b1 = Button(bg_img,command=self.facerecognize(),image=self.std_img1,cursor="hand2")
        std_b1.place(x=600,y=170,width=180,height=180)

        std_b1_1 = Button(bg_img,command=self.facerecognize(),text="Face Detector",cursor="hand2",font=("tahoma",15,"bold"),bg="white",fg="navyblue")
        std_b1_1.place(x=600,y=350,width=180,height=45)
        #-----------------------------------------------------------------------------------
        #saving button

        std_img_btn1=Image.open(r"Images_GUI\save-icon.png")
        std_img_btn1=std_img_btn1.resize((180,180),Image.ANTIALIAS)
        self.std_img2=ImageTk.PhotoImage(std_img_btn1)

        std_b2 = Button(bg_img,command=self.txttolist,image=self.std_img2,cursor="hand2")
        std_b2.place(x=800,y=170,width=180,height=180)

        std_b1_2 = Button(bg_img,command=self.txttolist,text="Save Attendance",cursor="hand2",font=("tahoma",15,"bold"),bg="white",fg="navyblue")
        std_b1_2.place(x=800,y=350,width=180,height=45)
    #=====================Attendance===================
    # def mark_attendance(self,n):
    #     column=0
    #     today = date.today()
    #     td=str(today)
    #     datedata=""
    #     datad=0
    #     column=0
    #     today = date.today()
    #     td=str(today)
    #     wb=load_workbook('sheet.xlsx')
    #     ws=wb.active
    #     sheets = wb.sheetnames
    #     ws = wb[sheets[0]]
    #     dateindex=0
    #     for dtrows in range(8,9):
    #         for dtcol in range(3,11):
    #             char=get_column_letter(dtcol)
    #             val=ws[char+str(dtrows)].value
    #             val=str(val)
    #             datedata=val[0:10]
    #             if td==datedata:
    #                 dateindex=dtcol
    #                 break
    #     if td==datedata:
    #         column=dateindex
    #     else:
    #         print("Date Error")
    #     rowdata=0
    #     for rows in range(10,41):
    #         for col in range(2,3):
    #             char=get_column_letter(col)
    #             val=ws[char+str(rows)].value
    #             for i in n:
    #                 if n==val:
    #                     row=rowdata+rows
    #                     #print(row)
    #                     c1 = ws.cell(row, column)
    #                     c1.value = "P" 
    #     wb.save("C:/Users/AMMAR MEER/Desktop/Python-FYP-Face-Recognition-Attendence-System-master/Attendance/sheet3.xlsx")

    def save_raw_Attendance(self,n):
        today = str(date.today())
        txt="raw_"+today+".txt"
        filepath = directory
        file=filepath+txt
        with open(filepath+"raw_"+today+".txt", "r+",newline="\n") as fp:
                myDatalist=fp.readlines()
                name_list=[]
                for line in myDatalist:
                    entry=line.split((","))
                    name_list.append(entry[0])
                    ##print(name_list)
                if((n not in name_list)):
                    now=datetime.now()
                    d1=now.strftime("%d/%m/%Y")
                    dtString=now.strftime("%H:%M:%S")
                    fp.writelines(n+ '\n')
        def remove_repeated(file):
            openFile = open(file, "r") 
            writeFile = open("text.txt", "w") 
            #Store traversed lines
            tmp = set() 
            for txtLine in openFile: 
            #Check new line
                if txtLine not in tmp: 
                    writeFile.write(txtLine) 
            #Add new traversed line to tmp 
                    tmp.add(txtLine)         
            openFile.close() 
            writeFile.close()
        remove_repeated(file)
                # with open("attendance.csv","r+") as f:
        #     # with open(filepath,'w+') as fp:
        #         myDatalist=f.readlines()
        #         name_list=[]
                    
        #         for line in myDatalist:
        #             entry=line.split((","))
        #             name_list.append(entry[0])
        #             print(name_list)

        #             if((i not in name_list)) and ((r not in name_list)) and ((n not in name_list)):
        #                     now=datetime.now()
        #                     d1=now.strftime("%d/%m/%Y")
        #                     dtString=now.strftime("%H:%M:%S")
        #                     f.writelines(f"\n{i},{r}, {n}, {dtString}, {d1}, Present")
        #         df=pd.read_csv("studentDetails.csv")
        #         col_names =  ['Id','Roll''Name','Time','Date','Status']
        #         attendance = pd.DataFrame(columns = col_names)
        #         attendance=attendance.drop_duplicates(subset=['Id'],keep='first')
        #         ts = time.time()      
        #         date = datetime.datetime.fromtimestamp(ts).strftime('%Y-%m-%d')
        #         timeStamp = datetime.datetime.fromtimestamp(ts).strftime('%H:%M:%S')
        #         Hour,Minute,Second=timeStamp.split(":")
        #         fileName="Attendance"+date+"_"+Hour+"-"+Minute+"-"+Second+".csv"
        #         attendance.to_csv(fileName,index=False)

    #================face recognition==================
    
    def markA(self,namelist):
        column=0
        today = date.today()
        td=str(today)
        datedata=""
        datad=0
        column=0
        today = date.today()
        td=str(today)
        wb=load_workbook('sheet.xlsx')
        ws=wb.active
        sheets = wb.sheetnames
        ws = wb[sheets[0]]
        dateindex=0
        for dtrows in range(8,9):
            for dtcol in range(3,11):
                char=get_column_letter(dtcol)
                val=ws[char+str(dtrows)].value
                val=str(val)
                datedata=val[0:10]
                if td==datedata:
                    dateindex=dtcol
                    break
        if td==datedata:
            column=dateindex
        else:
            print("Date Error")
        rowdata=0
        for rows in range(10,41):
            for col in range(2,3):
                char=get_column_letter(col)
                val=ws[char+str(rows)].value
                for i in namelist:
                    if i==val:
                        row=rowdata+rows
                        #print(row)
                        c1 = ws.cell(row, column)
                        c1.value = "P" 
        wb.save("C:/Users/AMMAR MEER/Desktop/Python-FYP-Face-Recognition-Attendence-System-master/Attendance/sheet3.xlsx")

    def txttolist(self):
        my_file = open("text.txt", "r")
        # reading the file
        data = my_file.read()
        # replacing end splitting the text 
        # when newline ('\n') is seen.
        data_into_list = data.split("\n")
        correctlist=data_into_list.pop()
        my_file.close()
        self.markA(data_into_list)

    def facerecognize(self):
            path = 'ImagesOfFaces'
            images = []
            classNames = []
            myList = os.listdir(path)
            print(myList)
            for cl in myList:
                curImg = cv2.imread(f'{path}/{cl}')
                images.append(curImg)
                classNames.append(os.path.splitext(cl)[0])
            print(classNames)
            def findEncodings(images):
                encodeList = []
                for img in images:
                    img = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)
                    encode = face_recognition.face_encodings(img)[0]
                    encodeList.append(encode)
                return encodeList
            encodeListKnown = findEncodings(images)
            print('Encoding Complete')
            cap = cv2.VideoCapture(0)
            while True:
                success, img = cap.read()
                #img = captureScreen()
                imgS = cv2.resize(img,(0,0),None,0.25,0.25)
                imgS = cv2.cvtColor(imgS, cv2.COLOR_BGR2RGB)
                facesCurFrame = face_recognition.face_locations(imgS)
                encodesCurFrame = face_recognition.face_encodings(imgS,facesCurFrame)
                for encodeFace,faceLoc in zip(encodesCurFrame,facesCurFrame):
                    matches = face_recognition.compare_faces(encodeListKnown,encodeFace)
                    faceDis = face_recognition.face_distance(encodeListKnown,encodeFace)
                    #print(faceDis)
                    matchIndex = np.argmin(faceDis)
                    if faceDis[matchIndex]< 0.50:
                        name = classNames[matchIndex]
                        #save_raw_Attendance(name) 
                        self.save_raw_Attendance(name)
                    else: name = 'Unknown'
                    #print(name)
                    y1,x2,y2,x1 = faceLoc
                    y1, x2, y2, x1 = y1*4,x2*4,y2*4,x1*4
                    cv2.rectangle(img,(x1,y1),(x2,y2),(0,0,255),2)
                    cv2.rectangle(img,(x1,y2-35),(x2,y2),(0,0,255),cv2.FILLED)
                    cv2.putText(img,name,(x1+6,y2-6),cv2.FONT_HERSHEY_COMPLEX,0.5,(255,255,255),1)
                    
                cv2.imshow('Webcam',img)
                if cv2.waitKey(10) == ord('q'):           
                    cap.release()
                    cv2.destroyAllWindows()

    # def face_recog(self):
        
    #     def draw_boundray(img,classifier,scaleFactor,minNeighbors,color,text,clf):
    #         gray_image=cv2.cvtColor(img,cv2.COLOR_BGR2GRAY)
    #         featuers=classifier.detectMultiScale(gray_image,scaleFactor,minNeighbors)
    #         coord=[]
    #         for (x,y,w,h) in featuers:
    #             cv2.rectangle(img,(x,y),(x+w,y+h),(0,255,0),3)
    #             id,predict=clf.predict(gray_image[y:y+h,x:x+w])
    #             confidence=int((100*(1-predict/300)))
    #             d=str(id)
    #             conn = mysql.connector.connect(host="localhost",username="root",password="SuperPC@467",database="face_recognition")
    #             cursor = conn.cursor()
    #             cursor.execute("select Name from add_student where Student_ID="+d)
    #             n=cursor.fetchone()
    #             n="+".join(n)
    #             cursor.execute("select Roll_No from add_student where Student_ID="+d)
    #             r=cursor.fetchone()
    #             r="+".join(r)
    #             cursor.execute("select Student_ID from add_student where Student_ID="+d)
    #             i=cursor.fetchone()
    #             i="+".join(i)
    #             cursor.execute("select Course from add_student where Student_ID="+d)
    #             c=cursor.fetchone()
    #             c="+".join(c)
    #             if confidence > 77:
    #                 cv2.putText(img,f"Student ID:{i}",(x,y-80),cv2.FONT_HERSHEY_COMPLEX,0.8,(64,15,223),2)
    #                 cv2.putText(img,f"Name:{n}",(x,y-55),cv2.FONT_HERSHEY_COMPLEX,0.8,(64,15,223),2)
    #                 cv2.putText(img,f"Roll-No:{r}",(x,y-30),cv2.FONT_HERSHEY_COMPLEX,0.8,(64,15,223),2)
    #                 cv2.putText(img,f"Accuracy:{confidence}",(x,y-10),cv2.FONT_HERSHEY_COMPLEX,0.8,(64,15,223),2)
    #                 self.save_raw_Attendance(n)                    
    #             else:
    #                 cv2.rectangle(img,(x,y),(x+w,y+h),(0,0,255),3)
    #                 cv2.putText(img,"Unknown Face",(x,y-5),cv2.FONT_HERSHEY_COMPLEX,0.8,(255,255,0),3)    
    #             coord=[x,y,w,y]
    #         return coord     
    #     #==========
    #     def recognize(img,clf,faceCascade):
    #         coord=draw_boundray(img,faceCascade,1.1,10,(255,25,255),"Face",clf)
    #         return img
    #     faceCascade=cv2.CascadeClassifier(haar)
    #     clf=cv2.face.LBPHFaceRecognizer_create()
    #     clf.read("clf.xml")
    #     videoCap=cv2.VideoCapture(0)
    #     while True:
    #         ret,img=videoCap.read()
    #         img=recognize(img,clf,faceCascade)
    #         cv2.imshow("Live Attendance",img)
    #         if cv2.waitKey(1) == 13:
    #             self.txttolist()
    #             break
    #     videoCap.release()
    #     cv2.destroyAllWindows()
    
if __name__ == "__main__":
    root=Tk()
    obj=FaceDetect(root)
    root.mainloop()